"""
Excel 报表导出工具
使用 openpyxl 生成带多 Sheet 页的 Excel 报表
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
from sqlalchemy.orm import Session
from collections import Counter

from app.models.inspection_record import InspectionRecord, ResultStatus
from app.models.device import Device
from app.models.user import User, UserRole


def generate_report_excel(records: list, start_date: str, end_date: str, db: Session) -> bytes:
    """
    生成包含多 Sheet 页的 Excel 报表
    - Sheet 1: 点检明细
    - Sheet 2: 异常汇总
    - Sheet 3: 人员达标率
    """
    wb = Workbook()

    # 获取设备和用户映射
    devices = {d.device_code: d.device_name for d in db.query(Device).all()}
    tech_users = {u.id: u.real_name for u in db.query(User).filter(User.role == UserRole.TECH).all()}

    # 样式定义
    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    def style_header(ws, headers, row=1):
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

    # ==================== Sheet 1: 点检明细 ====================
    ws1 = wb.active
    ws1.title = "点检明细"
    headers1 = ["记录ID", "设备编号", "设备名称", "技术员", "点检时间", "GPS纬度", "GPS经度",
                "结果状态", "异常原因", "工程师备注", "图片数量"]
    style_header(ws1, headers1)

    for row_idx, record in enumerate(records, 2):
        ws1.cell(row=row_idx, column=1, value=record.id)
        ws1.cell(row=row_idx, column=2, value=record.device_code)
        ws1.cell(row=row_idx, column=3, value=devices.get(record.device_code, ""))
        ws1.cell(row=row_idx, column=4, value=tech_users.get(record.tech_id, f"#{record.tech_id}"))
        ws1.cell(row=row_idx, column=5, value=record.check_time.strftime("%Y-%m-%d %H:%M:%S") if record.check_time else "")
        ws1.cell(row=row_idx, column=6, value=float(record.gps_lat) if record.gps_lat else "")
        ws1.cell(row=row_idx, column=7, value=float(record.gps_lng) if record.gps_lng else "")
        ws1.cell(row=row_idx, column=8, value="合格" if record.result_status == ResultStatus.PASS else "异常")
        ws1.cell(row=row_idx, column=9, value=record.remark or "")
        ws1.cell(row=row_idx, column=10, value=record.engineer_remark or "")
        ws1.cell(row=row_idx, column=11, value=len(record.photo_urls) if record.photo_urls else 0)

        # 异常行标黄
        if record.result_status == ResultStatus.FAIL:
            red_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            for col in range(1, 12):
                ws1.cell(row=row_idx, column=col).fill = red_fill

    # 设置列宽
    for col_letter, width in [("A", 10), ("B", 22), ("C", 18), ("D", 12), ("E", 20),
                                ("F", 14), ("G", 14), ("H", 10), ("I", 30), ("J", 30), ("K", 10)]:
        ws1.column_dimensions[col_letter].width = width

    # ==================== Sheet 2: 异常汇总 ====================
    ws2 = wb.create_sheet("异常汇总")
    headers2 = ["设备编号", "设备名称", "异常次数", "最近异常时间", "最近异常原因", "工程师处理状态"]
    style_header(ws2, headers2)

    fail_records = [r for r in records if r.result_status == ResultStatus.FAIL]
    device_fail_count = Counter(r.device_code for r in fail_records)
    device_last_fail = {}
    for r in fail_records:
        if r.device_code not in device_last_fail or r.check_time > device_last_fail[r.device_code]["check_time"]:
            device_last_fail[r.device_code] = {
                "check_time": r.check_time,
                "remark": r.remark,
                "engineer_remark": r.engineer_remark
            }

    for row_idx, (device_code, count) in enumerate(device_fail_count.most_common(), 2):
        last = device_last_fail.get(device_code, {})
        ws2.cell(row=row_idx, column=1, value=device_code)
        ws2.cell(row=row_idx, column=2, value=devices.get(device_code, ""))
        ws2.cell(row=row_idx, column=3, value=count)
        ws2.cell(row=row_idx, column=4, value=last.get("check_time", "").strftime("%Y-%m-%d %H:%M:%S") if last.get("check_time") else "")
        ws2.cell(row=row_idx, column=5, value=last.get("remark", ""))
        ws2.cell(row=row_idx, column=6, value="已处理" if last.get("engineer_remark") else "待处理")

    for col_letter, width in [("A", 22), ("B", 18), ("C", 10), ("D", 20), ("E", 30), ("F", 14)]:
        ws2.column_dimensions[col_letter].width = width

    # ==================== Sheet 3: 人员达标率 ====================
    ws3 = wb.create_sheet("人员达标率")
    headers3 = ["技术员", "点检总次数", "合格次数", "异常次数", "达标率"]
    style_header(ws3, headers3)

    tech_stats = {}
    for r in records:
        tech_id = r.tech_id
        if tech_id not in tech_stats:
            tech_stats[tech_id] = {"total": 0, "pass": 0, "fail": 0}
        tech_stats[tech_id]["total"] += 1
        if r.result_status == ResultStatus.PASS:
            tech_stats[tech_id]["pass"] += 1
        else:
            tech_stats[tech_id]["fail"] += 1

    for row_idx, (tech_id, stats) in enumerate(tech_stats.items(), 2):
        pass_rate = round(stats["pass"] / stats["total"] * 100, 1) if stats["total"] > 0 else 0
        ws3.cell(row=row_idx, column=1, value=tech_users.get(tech_id, f"#{tech_id}"))
        ws3.cell(row=row_idx, column=2, value=stats["total"])
        ws3.cell(row=row_idx, column=3, value=stats["pass"])
        ws3.cell(row=row_idx, column=4, value=stats["fail"])
        ws3.cell(row=row_idx, column=5, value=f"{pass_rate}%")

    for col_letter, width in [("A", 12), ("B", 12), ("C", 10), ("D", 10), ("E", 10)]:
        ws3.column_dimensions[col_letter].width = width

    # 保存到 BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()